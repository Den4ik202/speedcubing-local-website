import os
from openpyxl import Workbook, load_workbook
import re

DB_FILE = 'speedcubing_db.xlsx'


def init_db():
    """Инициализация базы данных: создание файла и листов, если их нет"""
    if not os.path.exists(DB_FILE):
        wb = Workbook()

        # Лист Участники
        ws_part = wb.active
        ws_part.title = "Участники"
        ws_part.append(["ID", "ФИО", "Дата рождения", "Прочее"])

        # Лист Соревнования
        wb.create_sheet("Соревнования")
        wb["Соревнования"].append(
            ["Comp_ID", "Название", "Адрес", "Дата начала", "Дата конца", "Информация", "Дисциплины"])

        # Лист Регистрации (связь участников и соревнований)
        wb.create_sheet("Регистрации")
        wb["Регистрации"].append(["Comp_ID", "ID_Участника"])

        # Лист Результаты
        wb.create_sheet("Результаты")
        wb["Результаты"].append(
            ["Comp_ID", "Дисциплина", "ID_Участника", "Попытка 1", "Попытка 2", "Попытка 3", "Попытка 4", "Попытка 5",
             "SINGLE", "AVG5"])

        wb.save(DB_FILE)


def get_next_participant_id():
    """Генерация следующего ID для участника (U-0001, U-0002 и т.д.)"""
    wb = load_workbook(DB_FILE)
    sheet = wb["Участники"]
    max_id = 0
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).startswith("U-"):
            try:
                num = int(str(row[0]).split("-")[1])
                if num > max_id:
                    max_id = num
            except ValueError:
                continue
    return f"U-{max_id + 1:04d}"


def search_participants(query):
    """Поиск участников по совпадению подстроки в ФИО (регистронезависимый)"""
    wb = load_workbook(DB_FILE)
    sheet = wb["Участники"]
    results = []
    q = query.lower().strip()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] and q in str(row[1]).lower():
            results.append({
                "id": row[0],
                "fio": row[1],
                "dob": row[2] if row[2] else "",
                "other": row[3] if row[3] else ""
            })
    return results


def save_participant_db(p_id, fio, dob, other):
    """Сохранение нового или обновление существующего участника"""
    wb = load_workbook(DB_FILE)
    sheet = wb["Участники"]
    updated = False

    # Если ID передан, пробуем найти и обновить
    if p_id:
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == p_id:
                row[1].value = fio
                row[2].value = dob
                row[3].value = other
                updated = True
                break

    # Если участник новый (не обновился или ID пустой)
    if not updated:
        p_id = get_next_participant_id()
        sheet.append([p_id, fio, dob, other])

    wb.save(DB_FILE)
    return p_id


def get_next_comp_id():
    """Генерация ID соревнования (C-0001, C-0002)"""
    wb = load_workbook(DB_FILE)
    sheet = wb["Соревнования"]
    max_id = 0
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).startswith("C-"):
            try:
                num = int(str(row[0]).split("-")[1])
                if num > max_id:
                    max_id = num
            except ValueError:
                continue
    return f"C-{max_id + 1:04d}"


def get_all_competitions():
    """Получить список всех соревнований"""
    wb = load_workbook(DB_FILE)
    sheet = wb["Соревнования"]
    comps = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            comps.append({
                "id": row[0],
                "name": row[1],
                "address": row[2],
                "start": row[3],
                "end": row[4],
                "info": row[5],
                "disciplines": row[6]
            })
    return comps


def get_competition(comp_id):
    """Получить данные одного соревнования по ID"""
    for comp in get_all_competitions():
        if comp["id"] == comp_id:
            return comp
    return None


def save_competition(comp_id, name, address, start, end, info, disciplines):
    """Сохранение или обновление соревнования"""
    wb = load_workbook(DB_FILE)
    sheet = wb["Соревнования"]
    updated = False

    # Строка с дисциплинами (через запятую)
    disc_str = ", ".join(disciplines) if isinstance(disciplines, list) else disciplines

    if comp_id:
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == comp_id:
                row[1].value = name
                row[2].value = address
                row[3].value = start
                row[4].value = end
                row[5].value = info
                row[6].value = disc_str
                updated = True
                break

    if not updated:
        comp_id = get_next_comp_id()
        sheet.append([comp_id, name, address, start, end, info, disc_str])

    wb.save(DB_FILE)
    return comp_id


def get_competition_participants(comp_id):
    """Получить список участников, зарегистрированных на конкретное соревнование"""
    wb = load_workbook(DB_FILE)
    sheet_reg = wb["Регистрации"]
    sheet_part = wb["Участники"]

    # Собираем ID участников этого соревнования
    part_ids = []
    for row in sheet_reg.iter_rows(min_row=2, values_only=True):
        if row[0] == comp_id:
            part_ids.append(row[1])

    # Получаем ФИО этих участников из листа Участники
    participants = []
    for row in sheet_part.iter_rows(min_row=2, values_only=True):
        if row[0] in part_ids:
            participants.append({
                "id": row[0],
                "fio": row[1]
            })
    return participants


def add_participant_to_comp(comp_id, part_id):
    """Добавить участника в соревнование"""
    wb = load_workbook(DB_FILE)
    sheet_reg = wb["Регистрации"]

    # Проверяем, нет ли его уже в списке (чтобы не добавить дважды)
    for row in sheet_reg.iter_rows(min_row=2, values_only=True):
        if row[0] == comp_id and row[1] == part_id:
            return False

    sheet_reg.append([comp_id, part_id])
    wb.save(DB_FILE)
    return True


def remove_participant_from_comp(comp_id, part_id):
    """Удалить участника из соревнования"""
    wb = load_workbook(DB_FILE)
    sheet_reg = wb["Регистрации"]

    # Ищем нужную строку и удаляем
    for idx, row in enumerate(sheet_reg.iter_rows(min_row=2), start=2):
        if row[0].value == comp_id and row[1].value == part_id:
            sheet_reg.delete_rows(idx)
            wb.save(DB_FILE)
            return True
    return False


def time_to_ms(t_str):
    """Преобразует строку времени WCA в миллисекунды для сортировки и вычислений"""
    if not t_str:
        return -1  # Пустое поле
    if t_str in ['DNF', 'DNS']:
        return float('inf')

    match = re.search(r'(\d+)\s*м\s*(\d+)\s*с\s*(\d+)\s*мс', t_str)
    if match:
        return int(match.group(1)) * 60000 + int(match.group(2)) * 1000 + int(match.group(3))
    return -1


def ms_to_time(ms):
    """Преобразует миллисекунды обратно в строку 'X1 м X2 с X3 мс'"""
    if ms == float('inf'):
        return "DNF"
    mins = int(ms // 60000)
    secs = int((ms % 60000) // 1000)
    mils = int(ms % 1000)
    return f"{mins} м {secs} с {mils:03d} мс"


def calc_stats(attempts):
    """Вычисление SINGLE и AVG5 по правилам WCA"""
    times = [time_to_ms(a) for a in attempts]

    # Считаем SINGLE (Лучшая из 5, исключая DNF/DNS)
    valid_times = [t for t in times if t != float('inf') and t != -1]
    single = ""
    if valid_times:
        single = ms_to_time(min(valid_times))
    elif any(a == 'DNF' for a in attempts if a):
        single = "DNF"
    elif any(a == 'DNS' for a in attempts if a):
        single = "DNS"

    # Считаем AVG5 (только если введены все 5 результатов)
    avg5 = ""
    filled_times = [t for t in times if t != -1]
    if len(filled_times) == 5:
        filled_times.sort()
        middle_3 = filled_times[1:4]  # Убираем лучшую [0] и худшую [4] попытки
        if float('inf') in middle_3:
            avg5 = "DNF"
        else:
            avg5 = ms_to_time(sum(middle_3) / 3)

    return single, avg5


def save_discipline_results(comp_id, discipline, part_id, attempts):
    """Сохранение результатов участника в базу"""
    wb = load_workbook(DB_FILE)
    sheet = wb["Результаты"]

    # Гарантируем, что список состоит ровно из 5 элементов
    attempts = (attempts + ["", "", "", "", ""])[:5]
    single, avg5 = calc_stats(attempts)

    updated = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == comp_id and row[1].value == discipline and row[2].value == part_id:
            for i in range(5):
                row[3 + i].value = attempts[i]
            row[8].value = single
            row[9].value = avg5
            updated = True
            break

    if not updated:
        sheet.append([comp_id, discipline, part_id] + attempts + [single, avg5])

    wb.save(DB_FILE)
    return True


def get_discipline_leaderboard(comp_id, discipline):
    """Получение отсортированной таблицы лидеров"""
    wb = load_workbook(DB_FILE)
    sheet_res = wb["Результаты"]
    sheet_part = wb["Участники"]

    fio_map = {row[0]: row[1] for row in sheet_part.iter_rows(min_row=2, values_only=True)}

    results = []
    for row in sheet_res.iter_rows(min_row=2, values_only=True):
        if row[0] == comp_id and row[1] == discipline:
            results.append({
                "part_id": row[2],
                "fio": fio_map.get(row[2], "Неизвестно"),
                "attempts": list(row[3:8]),
                "single": row[8] if row[8] else "",
                "avg5": row[9] if row[9] else "",
                "avg_ms": time_to_ms(row[9]),
                "single_ms": time_to_ms(row[8])
            })

    # Сортируем: сначала по AVG5 (меньше - лучше), затем по SINGLE
    # Если результата AVG5 еще нет (-1), убираем его в конец таблицы
    results.sort(key=lambda x: (
        x["avg_ms"] if x["avg_ms"] != -1 else float('inf'),
        x["single_ms"] if x["single_ms"] != -1 else float('inf')
    ))

    for i, res in enumerate(results):
        res["rank"] = i + 1

    return results


def get_participant_results(comp_id, discipline, part_id):
    """Поиск уже внесенных результатов участника (чтобы подгрузить их в поля)"""
    wb = load_workbook(DB_FILE)
    for row in wb["Результаты"].iter_rows(min_row=2, values_only=True):
        if row[0] == comp_id and row[1] == discipline and row[2] == part_id:
            return list(row[3:8])
    return ["", "", "", "", ""]